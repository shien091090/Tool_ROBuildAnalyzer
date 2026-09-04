"""一次性轉換腳本: 讀取使用者維護的《RO強化表.xlsx》, 產出 userdata/refine_rules.json
與 userdata/prices.json — 之後成本引擎以json為權威資料來源, 不再讀excel。

用法:
    py -3.12-64 scripts/convert_refine_rules.py [--excel PATH]

預設excel路徑: C:\\Users\\lithoshu\\Desktop\\強化表\\RO強化表.xlsx
輸出: <repo根目錄>/userdata/refine_rules.json, <repo根目錄>/userdata/prices.json
"""
from __future__ import annotations

import argparse
import json
from decimal import Decimal
from pathlib import Path

import openpyxl

DEFAULT_EXCEL_PATH = r"C:\Users\lithoshu\Desktop\強化表\RO強化表.xlsx"

# 材料正名對照 (excel原始資料存在的錯字/舊稱 -> 正名)
NAME_FIXES = {
    "乙太粉塵": "乙太星塵",  # spec §7.2規則9
    "鈽鐳": "鈰鐳",  # spec §12假設3 (含"鈽鐳礦石"→"鈰鐳礦石"、"乙太鈽鐳礦石"→"乙太鈰鐳礦石")
}


def fix_name(name: str) -> str:
    """套用材料正名規則(逐一取代已知錯字子字串)。"""
    for wrong, right in NAME_FIXES.items():
        name = name.replace(wrong, right)
    return name


def parse_rate(value) -> str:
    """把機率欄(可能是'90%'字串、'0.7'/0.7十進位、int)轉成十進位字串, 不用浮點誤差。"""
    text = str(value).strip()
    if text.endswith("%"):
        num = Decimal(text[:-1]) / Decimal(100)
    else:
        num = Decimal(text)
    num = num.normalize()
    out = format(num, "f")
    if out.startswith("."):
        out = "0" + out
    if out.startswith("-."):
        out = "-0" + out[1:]
    return out


def parse_int_zeny(value) -> int:
    """Zeny欄可能是int, 也可能是帶千分位逗號的字串(如'500,000')。"""
    if isinstance(value, (int, float)):
        return int(value)
    return int(str(value).replace(",", "").strip())


def parse_qty(value) -> int:
    """數量欄可能是int, 也可能是文字字串('1')。"""
    return int(str(value).strip())


def parse_fail_text(text: str) -> tuple[str, int]:
    """把防爆&防退/失敗欄文字轉成(fail, blessing)。

    對應規則(binding, 見task-1-brief.md):
      '-' -> safe
      '失敗時裝備消失'/'裝備消失' -> break
      '鐵匠的祝福xN' -> stay, blessing=N
      '失敗時裝備不消失'/'裝備不消失' -> stay, blessing=0
      '精煉度 -1'/'精煉度-1' -> minus1
    """
    text = text.strip()
    if text == "-":
        return "safe", 0
    if "鐵匠的祝福" in text:
        n = int(text.split("x")[-1].strip())
        return "stay", n
    if text in ("失敗時裝備消失", "裝備消失"):
        return "break", 0
    if text in ("失敗時裝備不消失", "裝備不消失"):
        return "stay", 0
    if "精煉度" in text and "-1" in text:
        return "minus1", 0
    raise ValueError(f"未知的防爆&防退/失敗欄文字: {text!r}")


def parse_ether_fail(fail_text: str, prevent_text: str) -> tuple[str, int]:
    """乙太表的失敗/防爆&防退為兩欄: 防爆&防退='不可使用'時無祝福可用, 該step的fail
    改看失敗欄(minus1/break/safe); 否則防爆&防退欄本身就是'鐵匠的祝福xN', 覆蓋為stay。
    """
    prevent_text = prevent_text.strip()
    if prevent_text == "不可使用":
        return parse_fail_text(fail_text)
    return parse_fail_text(prevent_text)


def parse_lv_range(text: str) -> tuple[int, int]:
    a, b = text.strip().split("~")
    return int(a), int(b)


def expand_steps(from_lv: int, to_lv: int, material: str, qty: int, rate: str,
                  fail: str, blessing: int) -> list[dict]:
    """把'14~18'這種範圍展開成逐級step, 各step參數相同。"""
    steps = []
    for lv in range(from_lv, to_lv):
        steps.append({
            "from": lv,
            "to": lv + 1,
            "material": material,
            "qty": qty,
            "rate": rate,
            "fail": fail,
            "blessing": blessing,
        })
    return steps


def parse_named_qty(text: str, sep: str = "x") -> tuple[str, int]:
    """'乙太天藍寶石 x 5' 或 '乙太魔石x3' 這種"名稱+分隔符+數量"格式解析。"""
    name_part, qty_part = text.rsplit(sep, 1)
    return fix_name(name_part.strip()), int(qty_part.strip())


def convert_refine_table(ws, display_col_b: str, display_col_c: str,
                          fail_mode: str) -> tuple[dict, dict]:
    """通用轉換: 精煉材料表/影子防具與手套/乙太防具與武器 三張表結構雷同
    (精煉度範圍/兩欄材料/共用數量+機率+失敗欄), 差異只在欄位順序與fail_mode。

    fail_mode:
      'single'  - 單一失敗欄(精煉材料表F欄、影子表G欄), 用parse_fail_text
      'ether'   - 失敗欄+防爆&防退欄兩欄合併判斷, 用parse_ether_fail
    """
    table_b = {"display": display_col_b, "steps": []}
    table_c = {"display": display_col_c, "steps": []}

    rows = list(ws.iter_rows(min_row=3, values_only=True))
    for row in rows:
        if row[0] is None:
            continue
        lv_from, lv_to = parse_lv_range(row[0])
        material_b = fix_name(str(row[1]).strip())
        material_c = fix_name(str(row[2]).strip())
        qty = parse_qty(row[3])

        if fail_mode == "single":
            # 精煉材料表: D數量,E機率,F防爆&防退 / 影子表: D數量,E花費(略),F機率,G失敗
            if ws.title == "精煉材料表":
                rate = parse_rate(row[4])
                fail, blessing = parse_fail_text(str(row[5]))
            else:  # 影子防具與手套
                rate = parse_rate(row[5])
                fail, blessing = parse_fail_text(str(row[6]))
        elif fail_mode == "ether":
            # 乙太防具與武器: D數量,E機率,F失敗,G防爆&防退
            rate = parse_rate(row[4])
            fail, blessing = parse_ether_fail(str(row[5]), str(row[6]))
        else:
            raise ValueError(f"未知fail_mode: {fail_mode}")

        table_b["steps"].extend(
            expand_steps(lv_from, lv_to, material_b, qty, rate, fail, blessing))
        table_c["steps"].extend(
            expand_steps(lv_from, lv_to, material_c, qty, rate, fail, blessing))

    return table_b, table_c


def convert_grade_steps(ws) -> list[dict]:
    steps = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        grade_from_raw, grade_to_raw = (s.strip() for s in row[0].split(">"))
        grade_from = "none" if grade_from_raw == "無" else grade_from_raw
        grade_to = grade_to_raw
        refine_req = int(row[1])
        rate = parse_rate(row[2])
        mat_name, mat_qty = parse_named_qty(str(row[3]))
        fee = parse_int_zeny(row[4])
        steps.append({
            "from": grade_from,
            "to": grade_to,
            "refine_req": refine_req,
            "rate": rate,
            "materials": [{"name": mat_name, "qty": mat_qty}],
            "fee": fee,
        })
    return steps


def convert_exchange_recipes(ws) -> dict:
    recipes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        name = fix_name(str(row[0]).strip())
        inputs = [parse_named_qty(part, sep="x")
                  for part in str(row[1]).split("、")]
        fee = parse_int_zeny(row[2])
        recipes[name] = {
            "inputs": [{"name": n, "qty": q} for n, q in inputs],
            "fee": fee,
        }
    return recipes


def convert_prices(ws) -> dict:
    prices = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        name = fix_name(str(row[0]).strip())
        prices[name] = parse_int_zeny(row[1])
    return prices


def convert(excel_path: Path) -> tuple[dict, dict]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    armor_lv1, weapon_lv4 = convert_refine_table(
        wb["精煉材料表"], "一級防具", "四級武器", fail_mode="single")
    shadow_armor, shadow_gauntlet = convert_refine_table(
        wb["影子防具與手套"], "影子防具", "影子手套", fail_mode="single")
    ether_armor2, ether_weapon5 = convert_refine_table(
        wb["乙太防具與武器"], "二級防具", "五級武器", fail_mode="ether")

    refine_rules = {
        "refine_tables": {
            "armor_lv1": armor_lv1,
            "weapon_lv4": weapon_lv4,
            "shadow_armor": shadow_armor,
            "shadow_gauntlet": shadow_gauntlet,
            "ether_armor2": ether_armor2,
            "ether_weapon5": ether_weapon5,
        },
        "blessing_item": "鐵匠的祝福",
        "grade_steps": convert_grade_steps(wb["升階機率表"]),
        "exchange_recipes": convert_exchange_recipes(wb["各材料兌換"]),
    }
    prices = convert_prices(wb["各材料價值"])
    return refine_rules, prices


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", default=DEFAULT_EXCEL_PATH,
                         help="來源excel路徑(預設為使用者桌面的強化表)")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    repo_root = Path(__file__).resolve().parent.parent
    userdata_dir = repo_root / "userdata"
    userdata_dir.mkdir(exist_ok=True)

    refine_rules, prices = convert(excel_path)

    refine_rules_path = userdata_dir / "refine_rules.json"
    prices_path = userdata_dir / "prices.json"
    refine_rules_path.write_text(
        json.dumps(refine_rules, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    prices_path.write_text(
        json.dumps(prices, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"已寫入 {refine_rules_path}")
    print(f"已寫入 {prices_path}")


if __name__ == "__main__":
    main()
